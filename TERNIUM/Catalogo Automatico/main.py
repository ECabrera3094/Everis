import time
import json
import openpyxl
import threading
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import logging

logging.basicConfig(level=logging.DEBUG)

def create_dfGroups():

    logging.info("Creacion del DF por Grupos: ")

    # Abrimos el Json para la Lectura de las Columnas del Grupo.
    with open("C:\\Users\\everis\\Documents\\TERNIUM\\Catalogo Automatico\\data\\data.json") as data:

        # Cargamos la Data del Json.
        data_JSON = json.loads(data.read())

        # Obtenemos el Diccionario del Json.
        data_Groups = data_JSON['dfGroups'][0]

        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        listValues = list(data_Groups.values())

        # Lista Final que usaremos para almacenar las Columnas del DF.
        listFinal_Values = [bytes(eachValue,'utf-8') for eachValue in listValues]

        # Obtenemos el Diccionario del Json para la Lectura de la Ruta del .DAT.
        DAT_File = data_JSON['DAT'][0]

        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        path_DAT_File = list(DAT_File.values())

        # La Elemento 0 sera la Ruta del .DAT
        path_DAT_File = path_DAT_File[0]    

    # Abrimos el .DAT
    with open(path_DAT_File, 'rb') as myDAT_File:

        # Lectura del .DAT
        myDAT_File = myDAT_File.readlines()

        # Creamos un DataFrame.
        df = pd.DataFrame()

        # Recorremos la Lista con las Palabras Clave que usaremos para las Columnas del DF.
        for i in range(len(listFinal_Values)):

            # Para Cada Linea del Archivo
            for eachLine in myDAT_File:
                
                # Validacion si la Palabra Clave aparece en el Archivo.
                if (listFinal_Values[i] in eachLine):
                    
                    try:
                        # Transformamos a UTF-8.
                        newString = eachLine.decode(encoding = "utf-8")
                    except:
                        # Transformamos a ISO-8859-1
                        newString = eachLine.decode(encoding = "ISO-8859-1")
                        
                    # Eliminamos el Tabulador y el Salto de Linea.
                    newString = newString.strip("\r\n")

                    # Buscamos los Elementos que estan Despues de los (:)
                    newString = newString.split(":")

                    nameColumn = listFinal_Values[i]

                    df = df.append( {nameColumn : newString[1]},  ignore_index=True)
                elif eachLine == b'endheader:\r\n':

                    break
        
        # DF Final que Usaremos
        dfGroups = pd.DataFrame() 

        for i in range(len(listFinal_Values)):

            # Nuevos DataFrames para Limpiar
            copy_df = pd.DataFrame() 

            copy_df = df[listFinal_Values[i]]

            copy_df = copy_df.dropna()

            copy_df = copy_df.reset_index(drop=True)

            dfGroups[listFinal_Values[i]] = copy_df

        # Obtenemos el Diccionario del Json para la Lectura de la Ruta del PKL del Grupo.
        pkl_File = data_JSON['txt_Data_Groups'][0]
        
        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        pkl_File = list(pkl_File.values())

        # La Elemento 0 sera la Ruta del PKL del Grupo.
        pkl_Path_File = pkl_File[0]

        dfGroups.to_pickle(pkl_Path_File)
        
        logging.info("FIN")

def create_dfSignals():
    
    logging.info("Creacion del DF por Signal: ")

    # Abrimos el Json.
    with open("C:\\Users\\everis\\Documents\\TERNIUM\\Catalogo Automatico\\data\\data.json") as data:

        # Cargamos la Data del Json.
        data_JSON = json.loads(data.read())

        # Obtenemos el Diccionario del Json.
        data_Signals = data_JSON['dfSignals'][0]

        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        listValues = list(data_Signals.values())

        # Lista Final que usaremos para almacenar las Columnas del DataFrame.
        # Convertimos cada Valor a Byte ya que Trabajaremos con el .DAT
        listFinal_Values = [bytes(eachValue,'utf-8') for eachValue in listValues]

        # Obtenemos el Diccionario del Json.
        DAT_File = data_JSON['DAT'][0]

        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        path_DAT_File = list(DAT_File.values())

        path_DAT_File = path_DAT_File[0] 

    # Abrimos el .DAT
    with open(path_DAT_File, 'rb') as myDAT_File:

        # Lectura del .DAT
        myDAT_File = myDAT_File.readlines()

        # Creamos un DataFrame.
        df = pd.DataFrame()

        # Bandera
        blnFlag = False

        # Recorremos la Lista con las Palabras Clave que usaremos para las Columnas del DF.
        for i in range(len(listFinal_Values)):

            # Para Cada Linea del Archivo
            for eachLine in myDAT_File:
                                 
                if (eachLine == b'endheader:\r\n'):
                    blnFlag = True

                # Si se Activo la Bandera
                if (blnFlag):

                    # Validacion si la Palabra Clave aparece en el Archivo.
                    if (listFinal_Values[i] in eachLine):
                        
                        try:
                            # Transformamos a UTF-8.
                            newString = eachLine.decode(encoding = "utf-8")
                        except:
                            # Transformamos a ISO-8859-1
                            newString = eachLine.decode(encoding = "ISO-8859-1")
                        
                        # Eliminamos PRIMERO posibles Espacios Vacios 
                        newString = newString.strip()

                        # Eliminamos el Tabulador y el Salto de Linea.
                        newString = newString.strip("\r\n")

                        # Eliminamos las Comillas en caso de que Existan al Principio y Fin.
                        newString = newString.lstrip('"')
                        
                        newString = newString.rstrip('"')

                        # Buscamos los Elementos que estan Despues de los (:)
                        # El Split (char, 1) SOLO dividira hasta la PRIMERA APARICION del Char.
                        newString = newString.split(":", 1)

                        nameColumn = listFinal_Values[i]

                        df = df.append( {nameColumn : newString[1]},  ignore_index=True)
                    elif eachLine == b'endASCII:\r\n':

                        break

            blnFlag = False

        # DF Final que Usaremos
        dfSignals = pd.DataFrame() 

        for i in range(len(listFinal_Values)):

            # Nuevos DataFrames para Limpiar
            copy_df = pd.DataFrame() 

            copy_df = df[listFinal_Values[i]]

            copy_df = copy_df.dropna()

            copy_df = copy_df.reset_index(drop = True)

            dfSignals[listFinal_Values[i]] = copy_df

        # Obtenemos el Diccionario del Json para la Lectura de la Ruta del PKL del Grupo.
        pkl_File = data_JSON['txt_Data_Signals'][0]
        
        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        pkl_File = list(pkl_File.values())

        # La Elemento 0 sera la Ruta del PKL del Grupo.
        pkl_Path_File = pkl_File[0]

        dfSignals.to_pickle(pkl_Path_File)
        
        logging.info("FIN")

def create_Catalogue():
    
    logging.info("Inicia creacion de Catalogo.")

    # Abrimos el Json.
    with open("C:\\Users\\everis\\Documents\\TERNIUM\\Catalogo Automatico\\data\\data.json") as data:

        # Abrimos los 2 DF: dfGroups y dfSignals (que se encuentran en Formatos PKL)
        data_JSON = json.loads(data.read())
        pkl_File = data_JSON['txt_Data_Groups'][0]
        pkl_File = list(pkl_File.values())
        Groups_Path_File = pkl_File[0]
        dfGroups = pd.read_pickle(Groups_Path_File)

        pkl_File = data_JSON['txt_Data_Signals'][0]
        pkl_File = list(pkl_File.values())
        Signals_Path_File = pkl_File[0]
        dfSignals = pd.read_pickle(Signals_Path_File)

        # Call a Workbook() function of openpyxl to create a new blank Workbook object.
        wb = openpyxl.Workbook()

        # Get workbook active sheet from the active attribute.
        sheet = wb.active

        sheet.title = "Catalogue"

        # ----- Creamos todos los Headers del Catalogo.
        columns_Catalogue = ['Index', 'Grupo', 'Nombre_Grupo', 'Nombre_Senial', 'Channel_Number', 'Unit', 'Nombre_IBA']
        
        for i in range(1, len(columns_Catalogue)):
            ## NOTA: El Excel empieza con la Celda (1,1)
            name_Columns_Catalogue = sheet.cell(row = 1, column = i)
            name_Columns_Catalogue.value = columns_Catalogue[i]

        # ----- Llenamos la Columna del Nombre_Grupo. -----
        # Lista donde alamcenaremos CADA UNO de los Elementos de (Name * SignalCount) en dfGroups.
        list_Nombre_Grupo = [] 

        for index, row in dfGroups.iterrows():
            for i in range( int(row[b'signalCount:']) ):
                list_Nombre_Grupo.append( row[b'name:'] )
                
        # Bandera que Inicializa en 2 porque en 1 esta el Header de la Columna.
        flag_row = 2

        # Iteramos CADA UNO de los Nombres del Grupo y los Insertamos en las Celdas HACIA ABAJO.
        for i in range(len(list_Nombre_Grupo)):
            value_Column_Nombre_Grupo = sheet.cell(row = flag_row, column = 2)
            value_Column_Nombre_Grupo.value = list_Nombre_Grupo[i]
            flag_row += 1

        # ----- Llenamos la Columna del Nombre_Senial
        # Lista donde alamcenaremos CADA UNO de los Elementos de (Name) en dfSignals.
        list_Nombre_Senial = [row[b'name:'] for index, row in dfSignals.iterrows()]

        # Bandera que Inicializa en 2 porque en 1 esta el Header de la Columna.
        flag_row = 2   

        # Iteramos CADA UNO de los Nombres del Grupo y los Insertamos en las Celdas HACIA ABAJO.
        for i in range(len(list_Nombre_Senial)):
            value_Column_Nombre_Senial = sheet.cell(row = flag_row, column = 3)
            value_Column_Nombre_Senial.value = list_Nombre_Senial[i]
            flag_row += 1

        # ----- Llenamos la Columna del Channel_Number.
        # Lista donde alamcenaremos CADA UNO de los Elementos de (beginchannel) en dfSignals.
        list_Channel_Number = [row[b'beginchannel:'] for index, row in dfSignals.iterrows()]

        # Bandera que Inicializa en 2 porque en 1 esta el Header de la Columna.
        flag_row = 2   

        for i in range(len(list_Channel_Number)):
            value_Column_Channel_Number = sheet.cell(row = flag_row, column = 4)
            value_Column_Channel_Number.value = int(list_Channel_Number[i])
            flag_row += 1

        # ----- Llenamos la Columna del Unit.
        # Lista donde alamcenaremos CADA UNO de los Elementos de (unit) en dfSignals.
        list_Unit = [row[b'unit:'] for index, row in dfSignals.iterrows()]

        # Bandera que Inicializa en 2 porque en 1 esta el Header de la Columna.
        flag_row = 2   
        
        for i in range(len(list_Unit)):
            value_Column_Unit = sheet.cell(row = flag_row, column = 5)
            value_Column_Unit.value = list_Unit[i]
            flag_row += 1

        # ----- Llenamos la Columna del Nombre_IBA
        # Es la Concatenacion del Nombre_Senial en Lowercase y el Channel_Number con 000
        flag_row = 2

        for i in range(len(list_Nombre_Senial)):
            # Eliminamos los Espacios y Sustituimos por (_). Convertimos a Minusculas.
            # Si el Nombre tiene DOBLE ESPACIO, se Sustituye por UNO.
            # Si la Lista tiene elemento, se trabaja aqui, sino Hasta que se Tenga un Nombre Correcto.
            if list_Nombre_Senial[i]:
                name_IBA = list_Nombre_Senial[i].replace(' ', '_').replace("  ", " ").replace("__", '_').replace('.', '_').replace(";", '_').replace("(", '_').replace(")", '_').replace('%', '_').replace(',', '_').replace('#', '_').replace('+', '_').replace('=', '_').replace(':', '_').replace('-', '_').replace('\'', '_').replace('^', '_').replace('?', '_').replace("'/'", '_').replace("{", '_').replace("}", '_').replace("[", '_').replace("]", '_').replace(">", '_').replace("<", '_').lower()

            if (int(list_Channel_Number[i]) <= 999):
                # Si es Menor a 999, se le asignan 0000.
                number_IBA = '%04d' % int(list_Channel_Number[i])
            else:
                number_IBA = int(list_Channel_Number[i])
            #Concatemoa el Nombre Final.
            complete_name_IBA = name_IBA + "_C" + str(number_IBA)
            value_Column_Nombre_IBA = sheet.cell(row = flag_row, column = 6)
            value_Column_Nombre_IBA.value = complete_name_IBA
            flag_row += 1

    
        # Obtenemos el Diccionario del Json.
        save_Path = data_JSON['Save Path'][0]
        
        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        save_Path = list(save_Path.values())

        # Tomamos la Primera Posicion de la Lista
        save_Path = save_Path[0]

        # Guardamos el Archivo.
        wb.save(save_Path)

        # Cerramos el Archivo
        wb.close()

        # Leemos el Catalogo en Formato DataFrame
        df_Data_Signals = pd.read_excel(save_Path)


        # Obtenemos el Diccionario del Json.
        INFO = data_JSON['INFO'][0]
        
        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        INFO_Path = list(INFO.values())

        INFO_Path = INFO_Path[0]

        # Leemos la INFO en Formato DataFrame
        df_INFO = pd.read_excel(INFO_Path)

   
        # Obtenemos el Diccionario del Json.
        Catalogue = data_JSON['CSV Catalogue Path'][0]

        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        CSV_Catalogue_Path = list(Catalogue.values())

        # Obtenemos la Ruta de la Lista.
        CSV_Catalogue_Path = CSV_Catalogue_Path[0]

        # Convertimos a CSV
        df_Data_Signals.to_csv(CSV_Catalogue_Path)

        # Leemos los CSV
        df_Data_Signals = pd.read_csv(CSV_Catalogue_Path, dtype = str)


        # Obtenemos el Diccionario del Json.
        INFO = data_JSON['CSV INFO Path'][0]
        
        # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
        INFO_Path = list(INFO.values())

        # Obtenemos la Ruta de la Lista.
        INFO_Path = INFO_Path[0]

        # Convertimos a CSV
        df_INFO.to_csv(INFO_Path)

        # Leemos los CSV
        df_INFO = pd.read_csv(INFO_Path, dtype = str)

    # Iniciamos la Creacion del Match.
    logging.info("\nInicia creacion del Match.")

    # Agregamos la Columna 'flag', llena de NaN, que Servira para Marcar las Celdas que YA Transcurrimos.
    df_INFO['flag'] = np.nan

    for i in range(len(df_Data_Signals.index)):

        for j in range(len(df_INFO.index)):
            
            # Validamos si la Columna flag de INFO esta Vacia.
            # Si esta VACIA, se valida la Senial.
            # Sino se pasa a la Siguiente Linea.
            myEmpty_cell = df_INFO.iloc[j]['flag']


            # SI esta Vacia la Bandera 'flag' y el Nombre de la Senial esta Vacio (NaN)
            # Las Celdas Vacias del CSV se consideran de Tipo Float64
            if  (np.isnan(myEmpty_cell)) and ( type(df_Data_Signals.iloc[i]['Nombre_Senial']) == float ):
                
                logging.info("V:" + str(i))

                #--- iloc para leer [Fila][Columna]
                #--- loc para escribiri [Fila, Columna]

                # 1.- Colocar una Bandera a INFO para que ya no cuente ese Match.
                df_INFO.loc[j, 'flag'] = 1

                # 2.- Debo Obtener el Valor del Grupo en INFO.
                var = df_INFO.iloc[j]['Grupo']

                # 3.- Debo Guardar el Match en el Catalogo
                # Eliminamos Posibles Textos que NO deban Ir.
                var = var.replace("_text", "")
                df_Data_Signals.loc[i, 'Grupo'] = var
                
                # 4.- Al NO tener Nombre, el Grupo pasa a ser el Nombre.
                # Nombre_Senial Debe ir sin Corchetes
                var = var.replace("[", "").replace("]", "").replace("_text", "")
                # Lo Escribimos en la Columna de 'Nombre_Senial'.
                df_Data_Signals.loc[i, 'Nombre_Senial'] = var

                # 5.- Al YA tener un Nombre, procedemos a crear el Nombre_IBA 
                name_IBA = var

                var_2 = df_Data_Signals.iloc[i]['Channel_Number']
                if (int(var_2) <= 999 ):
                    # Si es Menor a 999, se le asignan 0000.
                    number_IBA = '%04d' % int(var_2)
                    # Ya que andamos por Aqui, modificamos este mismo Numero a la Columna 'Channel_Number'
                    df_Data_Signals.loc[i, 'Channel_Number'] = number_IBA
                else:
                    number_IBA = int(var_2)

                #Concatemoa el Nombre Final.
                complete_name_IBA = name_IBA + "_C" + str(number_IBA)
                # Lo Escribimos en la Columna de 'Nombre_IBA'.
                df_Data_Signals.loc[i, 'Nombre_IBA'] = complete_name_IBA
                break
            else:
                pass

            # ------ SI esta Vacia la Bandera 'flag', PERO SI tenemos Nombre.
            if np.isnan(myEmpty_cell):
                
                # Volvemos a tratar el Problema de las "", tratando de Eliminarlos
                text_INFO = str(df_INFO.iloc[j]['Señal']).replace('"', '').strip()
                
                text_Sheet = str(df_Data_Signals.iloc[i]['Nombre_Senial']).replace('"', '').strip()

                # Comparamos los Nombres de la Senial y de INFO.
                if (text_INFO == text_Sheet):
                    
                    logging.info(">:" + str(i))

                    #--- iloc para leer [Fila][Columna]
                    #--- loc para escribiri [Fila, Columna]

                    # 1.- Colocar una Bandera a INFO para que ya no cuente ese Match.
                    df_INFO.loc[j, 'flag'] = 1

                    # 2.- Debo Obtener el Valor del Grupo en INFO.
                    var = df_INFO.iloc[j]['Grupo']

                    # 3.- Debo Guardar el Match en el Catlogo
                    # Eliminamos Posibles Textos que NO deban Ir.
                    var = var.replace("_text", "")

                    df_Data_Signals.loc[i, 'Grupo'] = var

                    # 4.- Modificamos el Nombre para AGREGARLE el PRIMER Numero del Grupo al Principio de este
                    # Obtengo el Indice que me Indica hasta que Cantidad de Caracteres termina el Primer Numero.
                    index_Grupo = var.find(":")

                    if not index_Grupo >= 0:
                        # Si no encuentra : que busque el .
                        index_Grupo = var.find(".")

                    # Delimito el Numero desde el Primer Caracter despues del '[' hasta mi Indice.
                    number_Grupo = var[1:index_Grupo]

                    # 5.- Obtengo el Nombre_Grupo
                    name = df_Data_Signals.iloc[i]['Nombre_Grupo']

                    # 6.- Concateno el Numero del Grupo con el Nombre.
                    final_name = str(number_Grupo) + ". " + str(name)

                    # 7.- Guardo de nueva cuenta el Nombre del Grupo en su respectiva Celda.
                    df_Data_Signals.loc[i, 'Nombre_Grupo'] = final_name
                    
                    # 8.- Modificamos este mismo Numero a la Columna 'Channel_Number' 
                    var_2 = df_Data_Signals.iloc[i]['Channel_Number']
                    if (int(var_2) <= 999 ):
                        # Si es Menor a 999, se le asignan 0000.
                        number_IBA = '%04d' % int(var_2)
                        # Ya que andamos por Aqui, modificamos este mismo Numero a la Columna 'Channel_Number'
                        df_Data_Signals.loc[i, 'Channel_Number'] = number_IBA

                    break
                else:
                    pass

    df_Data_Signals.to_excel("C:\\Users\\everis\\Documents\\TERNIUM\\Catalogo Automatico\\results\\Catalogue_2.xlsx")

    # ----- Obtenemos Ruta para Almacenar el Parquet.

    # Obtenemos el Diccionario del Json.
    parquet_Path = data_JSON['Parquet Path'][0]

    # Del Diccionario obtenemos los Valores y los transformamos a una Lista.
    parquet_Path = list(parquet_Path.values())

    # Convertirmos el DataFrame a Parquet y lo almacenamos en la Ruta Obtenida.
    df_Data_Signals.to_parquet(parquet_Path[0])

    # Fin del Match
    logging.info("FIN")


if __name__ == '__main__':
    
    start_time = time.time()

    dfGroups = threading.Thread(target=create_dfGroups)

    dfSignals = threading.Thread(target=create_dfSignals)

    dfGroups.start()
    dfSignals.start()

    dfGroups.join()
    dfSignals.join()
    
    create_Catalogue()
    
    print("--- %s seconds ---" % (time.time() - start_time))